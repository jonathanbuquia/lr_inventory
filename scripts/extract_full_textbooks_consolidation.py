import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ROOT = Path(
    r"C:\Users\Jonathan Buquia\OneDrive - Department of Education\SAMPLE CONSOLIDATED FOLDER"
)
OUTPUT_DIR = Path(
    r"C:\Users\Jonathan Buquia\Downloads\OFFICE\WEB\LR\outputs\full-consolidation-20260420"
)
OUTPUT_JSON = OUTPUT_DIR / "consolidation_dataset.json"
IGNORED_FILENAMES = {
    "grade1_consolidated_all_divisions.xlsx",
    "grade1_grade2_consolidated_all_divisions.xlsx",
    "consolidation.xlsx",
}


def collapse_spaces(value):
    return " ".join(str(value).strip().split())


def normalize_grade(raw_grade):
    if raw_grade is None:
        return None

    grade = collapse_spaces(raw_grade).upper()
    if not grade or grade in {"GRADE LEVEL", "GRADE"}:
        return None

    if grade == "KINDER":
        return "KINDER"

    if grade.isdigit():
        return f"GRADE {int(grade)}"

    match = re.fullmatch(r"SHS\s*[-(]?\s*(11|12)\)?", grade)
    if match:
        return f"GRADE {match.group(1)}"

    if grade == "SHS":
        return "SHS"

    return grade


def normalize_subject(raw_subject):
    if raw_subject is None:
        return None

    subject = collapse_spaces(raw_subject)
    if not subject or subject.upper() == "SUBJECTS":
        return None

    return subject.upper()


def to_number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return 0

    try:
        return int(float(text))
    except ValueError:
        return 0


def grade_sort_key(grade_label):
    if grade_label == "KINDER":
        return (0, 0, grade_label)

    if grade_label.startswith("GRADE "):
        suffix = grade_label.replace("GRADE ", "", 1)
        if suffix.isdigit():
            return (1, int(suffix), grade_label)

    if grade_label == "SHS":
        return (2, 0, grade_label)

    return (3, 0, grade_label)


def header_for(grade_label, subject_label):
    return f"{grade_label}: {subject_label}"


def extract_school(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "TextBooks" not in workbook.sheetnames:
        return None, set()

    worksheet = workbook["TextBooks"]
    row = {
        "DIVISION": path.parent.name,
        "SCHOOL NAME": path.stem,
    }
    seen_headers = set()

    for grade_raw, subject_raw, _, received_raw, *_rest in worksheet.iter_rows(
        min_row=2,
        max_col=8,
        values_only=True,
    ):
        grade_label = normalize_grade(grade_raw)
        subject_label = normalize_subject(subject_raw)
        if not grade_label or not subject_label:
            continue

        column_header = header_for(grade_label, subject_label)
        received = to_number(received_raw)

        row[column_header] = row.get(column_header, 0) + received
        seen_headers.add(column_header)

    return row, seen_headers


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    rows = []
    all_headers = set()
    skipped = []

    for path in sorted(SOURCE_ROOT.rglob("*.xlsx")):
        if path.name.lower() in IGNORED_FILENAMES:
            continue

        try:
            row, seen_headers = extract_school(path)
            if row is None:
                skipped.append({"file": str(path), "reason": "Missing TextBooks sheet"})
                continue

            rows.append(row)
            all_headers.update(seen_headers)
        except Exception as exc:
            skipped.append({"file": str(path), "reason": str(exc)})

    grouped_headers = defaultdict(list)
    for header in all_headers:
        grade_label, subject_label = header.split(": ", 1)
        grouped_headers[grade_label].append(subject_label)

    ordered_headers = ["DIVISION", "SCHOOL NAME"]
    for grade_label in sorted(grouped_headers, key=grade_sort_key):
        for subject_label in sorted(grouped_headers[grade_label]):
            ordered_headers.append(header_for(grade_label, subject_label))

    payload = {
        "source_root": str(SOURCE_ROOT),
        "record_count": len(rows),
        "skipped_count": len(skipped),
        "headers": ordered_headers,
        "rows": rows,
        "skipped": skipped,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} records and {len(ordered_headers)} headers to {OUTPUT_JSON}")
    print(f"Skipped {len(skipped)} files")


if __name__ == "__main__":
    main()
