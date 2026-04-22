import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ROOT = Path(
    r"C:\Users\Jonathan Buquia\OneDrive - Department of Education\SAMPLE CONSOLIDATED FOLDER"
)
OUTPUT_DIR = Path(
    r"C:\Users\Jonathan Buquia\Downloads\OFFICE\WEB\LR\outputs\grade1-consolidated-20260420"
)
OUTPUT_JSON = OUTPUT_DIR / "grade1_textbooks_consolidated.json"
IGNORED_FILENAMES = {
    "grade1_consolidated_all_divisions.xlsx",
    "grade1_grade2_consolidated_all_divisions.xlsx",
    "consolidated_all_divisions.xlsx",
}


SUBJECT_GROUPS = {
    "GRADE 1": {
        "GRADE 1: READING AND LITERACY": {
            "readingandliteracy",
            "readingandliteracyhiraya",
            "hiraya",
        },
        "GRADE 1: LANGUAGE": {
            "language",
            "languageiwika",
            "wika",
        },
        "GRADE 1: GMRC": {
            "gmrc",
            "gmrcwastongugalitamanggawi",
        },
        "GRADE 1: MAKABANSA": {
            "makabansa",
        },
        "GRADE 1: MATH": {
            "math",
        },
    },
    "GRADE 2": {
        "GRADE 2: READING AND LITERACY": {
            "english",
            "readingandliteracy",
        },
        "GRADE 2: LANGUAGE": {
            "bridgingprimerfilipino",
            "bridgingprimer2tagalogkontekstuwalisadongkagamitanngmagaaral",
            "bridgingprimertagalog",
            "filipino",
            "filipinotm",
            "filipinotx",
            "language",
        },
        "GRADE 2: GMRC": {
            "gmrc",
        },
        "GRADE 2: MAKABANSA": {
            "makabansa",
        },
        "GRADE 2: MATH": {
            "math",
        },
    },
}

ORDERED_HEADERS = [
    "DIVISION",
    "SCHOOL NAME",
    "GRADE 1: READING AND LITERACY",
    "GRADE 1: LANGUAGE",
    "GRADE 1: GMRC",
    "GRADE 1: MAKABANSA",
    "GRADE 1: MATH",
    "GRADE 2: READING AND LITERACY",
    "GRADE 2: LANGUAGE",
    "GRADE 2: GMRC",
    "GRADE 2: MAKABANSA",
    "GRADE 2: MATH",
]


def normalize(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def to_number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def match_group(grade_label, subject):
    normalized = normalize(subject)
    for title, variants in SUBJECT_GROUPS[grade_label].items():
        if normalized in variants:
            return title
    return None


def extract_school_file(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "TextBooks" not in workbook.sheetnames:
        return None

    worksheet = workbook["TextBooks"]
    row_data = {header: None for header in ORDERED_HEADERS}
    row_data["DIVISION"] = path.parent.name
    row_data["SCHOOL NAME"] = path.stem

    for values in worksheet.iter_rows(min_row=2, max_col=8, values_only=True):
        grade_level = values[0]
        subject = values[1]
        if not subject:
            continue

        grade_text = str(grade_level).strip()
        if grade_text not in {"1", "2"}:
            continue

        grade_label = f"GRADE {grade_text}"
        title = match_group(grade_label, subject)
        if not title:
            continue

        received = to_number(values[3])
        if row_data[title] is None:
            row_data[title] = received
        else:
            row_data[title] += received

    return row_data


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    rows = []
    skipped = []
    for path in sorted(SOURCE_ROOT.rglob("*.xlsx")):
        if path.name in IGNORED_FILENAMES:
            continue
        try:
            record = extract_school_file(path)
            if record is None:
                skipped.append({"file": str(path), "reason": "Missing TextBooks sheet"})
                continue
            rows.append(record)
        except Exception as exc:
            skipped.append({"file": str(path), "reason": str(exc)})

    payload = {
        "source_root": str(SOURCE_ROOT),
        "record_count": len(rows),
        "skipped_count": len(skipped),
        "rows": rows,
        "skipped": skipped,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} records to {OUTPUT_JSON}")
    if skipped:
        print(f"Skipped {len(skipped)} files")


if __name__ == "__main__":
    main()
