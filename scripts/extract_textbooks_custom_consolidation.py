import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ROOT = Path(
    r"C:\Users\Jonathan Buquia\OneDrive - Department of Education\SAMPLE CONSOLIDATED FOLDER"
)
OUTPUT_DIR = Path(
    r"C:\Users\Jonathan Buquia\Downloads\OFFICE\WEB\LR\outputs\consolidation-of-textbooks-20260421"
)
OUTPUT_JSON = OUTPUT_DIR / "consolidation_of_textbooks_dataset.json"
IGNORED_FILENAMES = {
    "grade1_consolidated_all_divisions.xlsx",
    "grade1_grade2_consolidated_all_divisions.xlsx",
    "consolidation.xlsx",
    "consolidation of textbooks.xlsx",
}

STANDARD_HEADERS = [
    "DIVISION",
    "SCHOOL NAME",
    "GRADE 1: Reading and Literacy",
    "GRADE 1: Language",
    "GRADE 1: GMRC",
    "GRADE 1: Makabansa",
    "GRADE 1: Math",
    "GRADE 2: Reading and Literacy",
    "GRADE 2: Language",
    "GRADE 2: GMRC",
    "GRADE 2: Makabansa",
    "GRADE 2: Math",
    "GRADE 3: Reading and Literacy",
    "GRADE 3: Language",
    "GRADE 3: GMRC",
    "GRADE 3: Makabansa",
    "GRADE 3: Math",
    "GRADE 3: Science",
    "GRADE 4: English",
    "GRADE 4: Filipino",
    "GRADE 4: Math",
    "GRADE 4: GMRC",
    "GRADE 4: Araling Panlipunan",
    "GRADE 4: Science",
    "GRADE 4: PE and Health",
    "GRADE 4: EPP",
    "GRADE 4: Music and Arts",
    "GRADE 5: English",
    "GRADE 5: Filipino",
    "GRADE 5: Math",
    "GRADE 5: GMRC",
    "GRADE 5: Araling Panlipunan",
    "GRADE 5: Science",
    "GRADE 5: PE and Health",
    "GRADE 5: EPP",
    "GRADE 5: Music and Arts",
    "GRADE 6: English",
    "GRADE 6: Filipino",
    "GRADE 6: Math",
    "GRADE 6: GMRC",
    "GRADE 6: Araling Panlipunan",
    "GRADE 6: Science",
    "GRADE 6: PE and Health",
    "GRADE 6: EPP",
    "GRADE 6: Music and Arts",
    "GRADE 7: English",
    "GRADE 7: Filipino",
    "GRADE 7: Math",
    "GRADE 7: VE",
    "GRADE 7: Araling Panlipunan",
    "GRADE 7: Science",
    "GRADE 7: PE and Health",
    "GRADE 7: TLE",
    "GRADE 7: Music and Arts",
    "GRADE 8: English",
    "GRADE 8: Filipino",
    "GRADE 8: Math",
    "GRADE 8: VE",
    "GRADE 8: Araling Panlipunan",
    "GRADE 8: Science",
    "GRADE 8: PE and Health",
    "GRADE 8: TLE",
    "GRADE 8: Music and Arts",
    "GRADE 9: English",
    "GRADE 9: Filipino",
    "GRADE 9: Math",
    "GRADE 9: VE",
    "GRADE 9: Araling Panlipunan",
    "GRADE 9: Science",
    "GRADE 9: PE and Health",
    "GRADE 9: TLE",
    "GRADE 9: Music and Arts",
    "GRADE 10: English",
    "GRADE 10: Filipino",
    "GRADE 10: Math",
    "GRADE 10: VE",
    "GRADE 10: Araling Panlipunan",
    "GRADE 10: Science",
    "GRADE 10: PE and Health",
    "GRADE 10: TLE",
    "GRADE 10: Music and Arts",
]

MAPPINGS = {
    ("GRADE 1", "GRADE 1: Reading and Literacy"): {
        "READING AND LITERACY",
        "HIRAYA",
        "READING AND LITERACY ( HIRAYA)",
        "READING AND LITERACY: HIRAYA",
    },
    ("GRADE 1", "GRADE 1: Language"): {
        "LANGUAGE",
        "LANGUAGE IWIKA)",
        "WIKA",
    },
    ("GRADE 1", "GRADE 1: GMRC"): {
        "GMRC",
        "GMRC: WASTONG UGALI, TAMANG GAWI",
    },
    ("GRADE 1", "GRADE 1: Makabansa"): {"MAKABANSA"},
    ("GRADE 1", "GRADE 1: Math"): {"MATH"},
    ("GRADE 2", "GRADE 2: Reading and Literacy"): {"READING AND LITERACY"},
    ("GRADE 2", "GRADE 2: Language"): {
        "LANGUAGE",
        "BRIDGING PRIMER - FILIPINO",
        "BRIDGING PRIMER 2 (TAGALOG) KONTEKSTUWALISADONG KAGAMITAN NG MAG-AARAL",
        "BRIDGING PRIMER TAGALOG",
        "ENGLISH",
        "FILIPINO",
        "FILIPINO TM",
        "FILIPINO TX",
    },
    ("GRADE 2", "GRADE 2: GMRC"): {"GMRC"},
    ("GRADE 2", "GRADE 2: Makabansa"): {"MAKABANSA"},
    ("GRADE 2", "GRADE 2: Math"): {"MATH"},
    ("GRADE 3", "GRADE 3: Reading and Literacy"): {"READING AND LITERACY"},
    ("GRADE 3", "GRADE 3: Language"): {"LANGUAGE", "ENGLISH", "FILIPINO"},
    ("GRADE 3", "GRADE 3: GMRC"): {"GMRC"},
    ("GRADE 3", "GRADE 3: Makabansa"): {"MAKABANSA"},
    ("GRADE 3", "GRADE 3: Math"): {"MATH"},
    ("GRADE 3", "GRADE 3: Science"): {"SCIENCE"},
    ("GRADE 4", "GRADE 4: English"): {"ENGLISH"},
    ("GRADE 4", "GRADE 4: Filipino"): {"FILIPINO", "FILIPINO TX"},
    ("GRADE 4", "GRADE 4: Math"): {"MATH"},
    ("GRADE 4", "GRADE 4: GMRC"): {"GMRC", "ESP"},
    ("GRADE 4", "GRADE 4: Araling Panlipunan"): {"ARALING PANLIPUNAN"},
    ("GRADE 4", "GRADE 4: Science"): {"SCIENCE"},
    ("GRADE 4", "GRADE 4: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 4", "GRADE 4: EPP"): {"E.P.P.", "EPP"},
    ("GRADE 4", "GRADE 4: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 5", "GRADE 5: English"): {"ENGLISH"},
    ("GRADE 5", "GRADE 5: Filipino"): {"FILIPINO"},
    ("GRADE 5", "GRADE 5: Math"): {"MATH"},
    ("GRADE 5", "GRADE 5: GMRC"): {"GMRC"},
    ("GRADE 5", "GRADE 5: Araling Panlipunan"): {"ARALING PANLIPUNAN"},
    ("GRADE 5", "GRADE 5: Science"): {"SCIENCE"},
    ("GRADE 5", "GRADE 5: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 5", "GRADE 5: EPP"): {"E.P.P.", "EPP"},
    ("GRADE 5", "GRADE 5: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 6", "GRADE 6: English"): {"ENGLISH"},
    ("GRADE 6", "GRADE 6: Filipino"): {"FILIPINO"},
    ("GRADE 6", "GRADE 6: Math"): {"MATH"},
    ("GRADE 6", "GRADE 6: GMRC"): {"GMRC"},
    ("GRADE 6", "GRADE 6: Araling Panlipunan"): {"ARALING PANLIPUNAN"},
    ("GRADE 6", "GRADE 6: Science"): {"SCIENCE"},
    ("GRADE 6", "GRADE 6: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 6", "GRADE 6: EPP"): {"E.P.P.", "EPP"},
    ("GRADE 6", "GRADE 6: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 7", "GRADE 7: English"): {"ENGLISH"},
    ("GRADE 7", "GRADE 7: Filipino"): {"FILIPINO"},
    ("GRADE 7", "GRADE 7: Math"): {"MATH"},
    ("GRADE 7", "GRADE 7: VE"): {"VE"},
    ("GRADE 7", "GRADE 7: Araling Panlipunan"): {"ARALING PANLIPUNAN", "AP"},
    ("GRADE 7", "GRADE 7: Science"): {"SCIENCE"},
    ("GRADE 7", "GRADE 7: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 7", "GRADE 7: TLE"): {"TLE"},
    ("GRADE 7", "GRADE 7: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 8", "GRADE 8: English"): {"ENGLISH"},
    ("GRADE 8", "GRADE 8: Filipino"): {"FILIPINO"},
    ("GRADE 8", "GRADE 8: Math"): {"MATH"},
    ("GRADE 8", "GRADE 8: VE"): {"VE"},
    ("GRADE 8", "GRADE 8: Araling Panlipunan"): {"ARALING PANLIPUNAN", "AP"},
    ("GRADE 8", "GRADE 8: Science"): {"SCIENCE"},
    ("GRADE 8", "GRADE 8: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 8", "GRADE 8: TLE"): {"TLE"},
    ("GRADE 8", "GRADE 8: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 9", "GRADE 9: English"): {"ENGLISH"},
    ("GRADE 9", "GRADE 9: Filipino"): {
        "FILIPINO",
        "FILIPINO (NOLI ME TANGERE)",
        "FILIPINO (NOLI)",
        "FILIPINO(NOLI ME TANGER)",
        "FILIPINO/NOLI ME TANGERE",
        "NOLI ME TANGERE",
    },
    ("GRADE 9", "GRADE 9: Math"): {"MATH"},
    ("GRADE 9", "GRADE 9: VE"): {"VE"},
    ("GRADE 9", "GRADE 9: Araling Panlipunan"): {"ARALING PANLIPUNAN", "AP"},
    ("GRADE 9", "GRADE 9: Science"): {"SCIENCE"},
    ("GRADE 9", "GRADE 9: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 9", "GRADE 9: TLE"): {"TLE"},
    ("GRADE 9", "GRADE 9: Music and Arts"): {"MUSIC AND ARTS"},
    ("GRADE 10", "GRADE 10: English"): {"ENGLISH"},
    ("GRADE 10", "GRADE 10: Filipino"): {
        "FILIPINO",
        "FILIPINO (EL FILI)",
        "FILIPINO (EL FILIBUSTERISMO)",
        "FILIPINO/EL FILIBUSTERISMO",
        "EL FILIBUSTERISMO",
    },
    ("GRADE 10", "GRADE 10: Math"): {"MATH"},
    ("GRADE 10", "GRADE 10: VE"): {"VE"},
    ("GRADE 10", "GRADE 10: Araling Panlipunan"): {"ARALING PANLIPUNAN", "AP"},
    ("GRADE 10", "GRADE 10: Science"): {"SCIENCE"},
    ("GRADE 10", "GRADE 10: PE and Health"): {"PE AND HEALTH"},
    ("GRADE 10", "GRADE 10: TLE"): {"TLE"},
    ("GRADE 10", "GRADE 10: Music and Arts"): {"MUSIC AND ARTS"},
}


def collapse_spaces(value):
    return " ".join(str(value).strip().split())


def normalize_grade(raw_grade):
    if raw_grade is None:
        return None

    grade = collapse_spaces(raw_grade).upper()
    if not grade or grade in {"GRADE LEVEL", "GRADE"}:
        return None

    if grade.isdigit():
        return f"GRADE {int(grade)}"

    return grade


def normalize_subject(raw_subject):
    if raw_subject is None:
        return None

    subject = collapse_spaces(raw_subject).upper()
    if not subject or subject == "SUBJECTS":
        return None

    return subject


def to_number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return 0

    try:
        return int(float(text))
    except ValueError:
        return 0


def build_subject_lookup():
    lookup = {}
    for (grade_label, header), subjects in MAPPINGS.items():
        for subject in subjects:
            lookup[(grade_label, subject)] = header
    return lookup


def extract_school(path, subject_lookup):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "TextBooks" not in workbook.sheetnames:
        return None

    row = {header: 0 for header in STANDARD_HEADERS}
    row["DIVISION"] = path.parent.name
    row["SCHOOL NAME"] = path.stem

    worksheet = workbook["TextBooks"]
    for grade_raw, subject_raw, _, received_raw, *_rest in worksheet.iter_rows(
        min_row=2,
        max_col=8,
        values_only=True,
    ):
        grade_label = normalize_grade(grade_raw)
        subject_label = normalize_subject(subject_raw)
        if not grade_label or not subject_label:
            continue

        header = subject_lookup.get((grade_label, subject_label))
        if not header:
            continue

        row[header] += to_number(received_raw)

    return row


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    subject_lookup = build_subject_lookup()

    rows = []
    skipped = []
    for path in sorted(SOURCE_ROOT.rglob("*.xlsx")):
        if path.name.lower() in IGNORED_FILENAMES:
            continue

        try:
            row = extract_school(path, subject_lookup)
            if row is None:
                skipped.append({"file": str(path), "reason": "Missing TextBooks sheet"})
                continue
            rows.append(row)
        except Exception as exc:
            skipped.append({"file": str(path), "reason": str(exc)})

    payload = {
        "source_root": str(SOURCE_ROOT),
        "record_count": len(rows),
        "skipped_count": len(skipped),
        "headers": STANDARD_HEADERS,
        "rows": rows,
        "skipped": skipped,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} records to {OUTPUT_JSON}")
    print(f"Skipped {len(skipped)} files")


if __name__ == "__main__":
    main()
